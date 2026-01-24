#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MDV Records Updater (World Aquatics + Sudamericano + PanAm Games)

- country/city: datos del atleta (según tu regla)
- competition_location: lugar del evento (texto)
- athlete_name: SOLO nombres
- type_probe: individual | relay
- Parche 23505 (duplicate key) + fix 'location' no inicializada
"""

from __future__ import annotations

import os
import re
import sys
import json
import uuid
import shutil
from dataclasses import dataclass
from datetime import datetime, timezone, date
from typing import Any, Dict, Iterable, List, Optional, Tuple

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
from supabase import create_client

# (Opcional) .env local
try:
    from dotenv import load_dotenv  # type: ignore
    load_dotenv()
except ModuleNotFoundError:
    pass


# -------------------------- Config --------------------------

MDV_UPDATER_VERSION = os.getenv("MDV_UPDATER_VERSION", "WA+SUDAM+PANAM+ARG_v15.5_STABLE_XLSX_MARKED_2026-01-24")
RUN_ID = os.getenv("GITHUB_RUN_ID") or str(uuid.uuid4())
RUN_TS = datetime.now(timezone.utc).isoformat(timespec="seconds")

# --- Self-check fingerprint (to confirm the runner is executing the expected file) ---
try:
    import hashlib
    _p = os.path.abspath(__file__)
    with open(_p, "rb") as _f:
        _sha = hashlib.sha256(_f.read()).hexdigest()[:16]
    print(f"🧩 SCRIPT_MARKER={MDV_UPDATER_VERSION}")
    print(f"🧩 SCRIPT_FILE={_p}")
    print(f"🧩 SCRIPT_SHA256_16={_sha}")
except Exception as _e:
    print(f"🧩 SCRIPT_FINGERPRINT_ERROR={_e}")
# -------------------------------------------------------------------------------
RUN_DATE = datetime.now(timezone.utc).date().isoformat()

SUPABASE_URL = os.getenv("SUPABASE_URL", "").strip()
SUPABASE_KEY = os.getenv("SUPABASE_KEY", "").strip()

HTTP_UA = os.getenv("MDV_HTTP_UA", "Mozilla/5.0 (MDV Records Updater)")


# Flags
MDV_STRICT = os.getenv("MDV_STRICT", "0").strip() == "1"
WA_INCLUDE_MIXED = os.getenv("WA_INCLUDE_MIXED", "0").strip() == "1"

# Fuentes Wiki (estables)
WIKI_SUDAM_URL = "https://es.wikipedia.org/wiki/Anexo:Plusmarcas_de_Sudam%C3%A9rica_de_nataci%C3%B3n"
WIKI_PANAM_GAMES_URL = "https://en.wikipedia.org/wiki/List_of_Pan_American_Games_records_in_swimming"
WIKI_ARG_URLS = [
    "https://es.wikipedia.org/wiki/R%C3%A9cords_argentinos_absolutos_de_nataci%C3%B3n",
    "https://en.wikipedia.org/wiki/List_of_Argentine_records_in_swimming",
]


# Fuentes ARG (estrategia)
SWIMCLOUD_ARG_RECORDS_URL = "https://www.swimcloud.com/country/arg/records/"
CADDA_RECORDS_DIR_URL = "https://cadda.org.ar/records/"
FENABA_RECORDS_URL = "https://fenaba.org.ar/records-argentinos/"


STROKE_MAP = {
    "freestyle": "Libre",
    "backstroke": "Espalda",
    "breaststroke": "Pecho",
    "butterfly": "Mariposa",
    "medley": "Combinado",
    "individual medley": "Combinado",
    "im": "Combinado",
    "medley relay": "Combinado",
    "freestyle relay": "Libre",
    "backstroke relay": "Espalda",
    "breaststroke relay": "Pecho",
    "butterfly relay": "Mariposa",
    "libre": "Libre",
    "espalda": "Espalda",
    "pecho": "Pecho",
    "mariposa": "Mariposa",
    "combinado": "Combinado",
}

# -------------------------- Helpers: time/date --------------------------

def _strip(s: Any) -> str:
    if s is None:
        return ""
    return str(s).strip()

def parse_time_to_ms(raw: Any) -> Optional[int]:
    if raw is None:
        return None
    s = _strip(raw)
    if not s:
        return None
    s = s.replace(",", ".")
    s = re.sub(r"\s+", "", s)

    if ":" in s:
        parts = s.split(":")
        try:
            if len(parts) == 3:
                hh = int(parts[0])
                mm = int(parts[1])
                sec = float(parts[2])
            elif len(parts) == 2:
                hh = 0
                mm = int(parts[0])
                sec = float(parts[1])
            else:
                return None
            return int(round(((hh * 3600 + mm * 60) + sec) * 1000))
        except Exception:
            return None

    dot_parts = s.split(".")
    try:
        if len(dot_parts) == 2:
            return int(round(float(s) * 1000))
        if len(dot_parts) == 3:
            mm = int(dot_parts[0]); ss = int(dot_parts[1]); cc = int(dot_parts[2])
            return (mm * 60 + ss) * 1000 + int(round(cc * 10))
        if len(dot_parts) == 4:
            hh = int(dot_parts[0]); mm = int(dot_parts[1]); ss = int(dot_parts[2]); cc = int(dot_parts[3])
            return (hh * 3600 + mm * 60 + ss) * 1000 + int(round(cc * 10))
    except Exception:
        return None
    return None

def format_ms_to_hms_2dp(ms: int) -> str:
    if ms < 0:
        ms = 0
    total_seconds = ms // 1000
    cent = (ms % 1000) // 10
    hh = total_seconds // 3600
    mm = (total_seconds % 3600) // 60
    ss = total_seconds % 60
    return f"{hh:02d}:{mm:02d}:{ss:02d}.{cent:02d}"

def parse_date(raw: Any) -> Optional[str]:
    """
    Normaliza fechas a ISO (YYYY-MM-DD) para insertar en columna DATE.
    - Limpia notas de Wikipedia tipo "[c]" o "[ note 1 ]"
    - Soporta formatos EN y ES comunes.
    """
    if raw is None:
        return None
    if isinstance(raw, (datetime, date)):
        return raw.date().isoformat() if isinstance(raw, datetime) else raw.isoformat()

    s = _strip(raw)
    if not s:
        return None

    # Quita referencias/footnotes tipo "[c]" / "[ note 1 ]"
    s = re.sub(r"\s*\[[^\]]+\]\s*", " ", s).strip()

    # Normaliza espacios
    s = re.sub(r"\s+", " ", s).strip()

    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return s

    # English full / abbreviated
    for fmt in ("%d %B %Y", "%d %b %Y", "%B %d, %Y", "%b %d, %Y"):
        try:
            dt = datetime.strptime(s, fmt)
            return dt.date().isoformat()
        except Exception:
            pass

    # Spanish: "21 de diciembre de 2023"
    m_es = re.match(r"^(\d{1,2})\s+de\s+([A-Za-zÁÉÍÓÚÜÑáéíóúüñ]+)\s+de\s+(\d{4})$", s, re.I)
    if m_es:
        d = int(m_es.group(1))
        mes = m_es.group(2).lower()
        y = int(m_es.group(3))
        meses = {
            "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
            "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9,
            "octubre": 10, "noviembre": 11, "diciembre": 12,
        }
        mo = meses.get(mes)
        if mo:
            try:
                return datetime(y, mo, d).date().isoformat()
            except Exception:
                return None

    # Numeric: 21/12/2023 or 21-12-23
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})$", s)
    if m:
        d = int(m.group(1)); mo = int(m.group(2)); y = int(m.group(3))
        if y < 100:
            y += 2000
        try:
            return datetime(y, mo, d).date().isoformat()
        except Exception:
            return None

    return None

def gender_label(g: Any) -> str:
    return "M" if _strip(g).upper().startswith("M") else "F"

def pool_label(pool: Any) -> str:
    p = _strip(pool).upper()
    if p in ("LCM", "50M", "50", "L", "LONG"):
        return "LCM"
    if p in ("SCM", "25M", "25", "S", "SHORT"):
        return "SCM"
    if p in ("SCY", "YARDS", "YD", "Y"):
        return "SCY"
    return p or "LCM"

# -------------------------- Event parsing (incluye relevos) --------------------------

RE_RELAY = re.compile(
    r"(?P<n>\d)\s*[x×]\s*(?P<leg>\d{2,4})\s*m\s*(?P<stroke>freestyle|backstroke|breaststroke|butterfly|medley)\s*(relay)?",
    re.IGNORECASE,
)
RE_INDIV = re.compile(
    r"(?P<dist>\d{2,4})\s*m\s*(?P<stroke>freestyle|backstroke|breaststroke|butterfly|medley|individual\s+medley|im|libre|espalda|pecho|mariposa|combinado)",
    re.IGNORECASE,
)

def parse_event(event_raw: Any) -> Tuple[Optional[int], Optional[str], str]:
    s = _strip(event_raw)
    if not s:
        return None, None, "individual"
    lo = s.lower()

    m = RE_RELAY.search(lo)
    if m:
        n = int(m.group("n"))
        leg = int(m.group("leg"))
        stroke_key = m.group("stroke").lower()
        stroke = STROKE_MAP.get(stroke_key)
        return n * leg, stroke, "relay"

    m = RE_INDIV.search(lo)
    if not m:
        return None, None, "individual"
    dist = int(m.group("dist"))
    stroke_key = m.group("stroke").lower().replace("  ", " ").strip()
    stroke = STROKE_MAP.get(stroke_key, STROKE_MAP.get(stroke_key.replace("individual ", "")))
    return dist, stroke, "individual"

# -------------------------- Supabase helpers --------------------------

def is_duplicate_error(e: Exception) -> bool:
    s = str(e).lower()
    return ("23505" in s) or ("duplicate key value" in s)

def _is_empty(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str):
        return v.strip() == ""
    return False

class SB:
    def __init__(self, url: str, key: str):
        if not url or not key:
            raise RuntimeError("Faltan SUPABASE_URL / SUPABASE_KEY")
        self.client = create_client(url, key)
        self.columns = self._detect_columns()
        print(f"🧬 DB columns detectadas: {len(self.columns)}")

    def _detect_columns(self) -> set:
        try:
            resp = self.client.table("records_standards").select("*").limit(1).execute()
            if resp.data:
                return set(resp.data[0].keys())
        except Exception:
            pass
        return {
            "id","gender","category","pool_length","stroke","distance",
            "time_clock","time_ms","time_clock_2dp",
            "record_scope","record_type",
            "competition_name","competition_location",
            "athlete_name","record_date",
            "city","country",
            "last_updated","source_url","source_name","source_note",
            "verified","updated_at","is_active","type_probe",
        }

    def _filter_payload(self, payload: Dict[str, Any], keep_empty: Iterable[str]) -> Dict[str, Any]:
        out: Dict[str, Any] = {}
        keep_empty_set = set(keep_empty)
        for k, v in payload.items():
            if k not in self.columns:
                continue
            if k in keep_empty_set:
                out[k] = v
                continue
            if _is_empty(v):
                continue
            out[k] = v
        return out

    def _fetch_existing(self, key: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        q = self.client.table("records_standards").select("*")
        for k, v in key.items():
            q = q.eq(k, v)
        resp = q.limit(1).execute()
        return (resp.data or [None])[0]

    def upsert_record(self, payload_full: Dict[str, Any]) -> str:
        key_fields = ["gender","category","pool_length","stroke","distance","record_type","record_scope"]
        key = {k: payload_full.get(k) for k in key_fields}
        for k in key_fields:
            if _is_empty(key.get(k)):
                raise ValueError(f"payload missing {k}")

        existing = self._fetch_existing(key)

        new_ms = payload_full.get("time_ms")
        old_ms = existing.get("time_ms") if existing else None
        time_changed = (existing is not None) and (new_ms is not None) and (old_ms is not None) and (int(new_ms) != int(old_ms))

        fill_fields = [
            "athlete_name",
            "country", "city",
            "record_date",
            "competition_name", "competition_location",
            "source_name", "source_url", "source_note",
            "type_probe",
        ]

        if existing:
            updates: Dict[str, Any] = {}

            for f in fill_fields:
                newv = payload_full.get(f)
                oldv = existing.get(f)
                if _is_empty(newv):
                    continue
                if _is_empty(oldv):
                    updates[f] = newv

            if time_changed:
                updates["time_ms"] = int(payload_full["time_ms"])
                updates["time_clock_2dp"] = payload_full.get("time_clock_2dp")
                updates["time_clock"] = payload_full.get("time_clock")

            if updates:
                updates["last_updated"] = RUN_DATE
                upd = self._filter_payload(updates, keep_empty=["last_updated"])
                self.client.table("records_standards").update(upd).eq("id", existing["id"]).execute()
                return "updated" if time_changed else "filled"

            return "unchanged"

        insert_payload = dict(payload_full)
        insert_payload["last_updated"] = RUN_DATE

        filtered = self._filter_payload(
            insert_payload,
            keep_empty=["gender","category","pool_length","stroke","distance","record_type","record_scope","type_probe","last_updated"]
        )
        try:
            self.client.table("records_standards").insert(filtered).execute()
            return "inserted"
        except Exception as e:
            if is_duplicate_error(e):
                # Re-fetch y pasar a modo update/fill
                existing2 = self._fetch_existing(key)
                if existing2:
                    return self.upsert_record(payload_full)
            raise

# -------------------------- World Aquatics --------------------------

@dataclass
class WASpec:
    code: str
    pool: str
    gender: str

def wa_url(spec: WASpec) -> str:
    base = "https://www.worldaquatics.com/swimming/records"
    if spec.code == "WR":
        return f"{base}?recordType=WR&eventTypeId=&region=&countryId=&gender={spec.gender}&pool={spec.pool}"
    if spec.code == "OR":
        return f"{base}?recordType=OR&eventTypeId=&region=&countryId=&gender={spec.gender}&pool={spec.pool}"
    if spec.code == "WJ":
        return f"{base}?recordCode=WJ&eventTypeId=&region=&countryId=&gender={spec.gender}&pool={spec.pool}"
    if spec.code == "CR_AMERICAS":
        return f"{base}?recordType=PAN&recordCode=CR&eventTypeId=&region=AMERICAS&countryId=&gender={spec.gender}&pool={spec.pool}"
    raise ValueError(spec.code)

def wa_scope_and_type(code: str, pool: str) -> Tuple[str, str]:
    pool = pool_label(pool)
    is_scm = (pool == "SCM")
    if code == "WR":
        return "Mundial", "Récord Mundial" + (" SC" if is_scm else "")
    if code == "OR":
        return "Olímpico", "Récord Olímpico"
    if code == "WJ":
        return "Mundial", "Récord Mundial Junior" + (" SC" if is_scm else "")
    if code == "CR_AMERICAS":
        return "Américas", "Récord Continental Américas" + (" SC" if is_scm else "")
    raise ValueError(code)

def wa_specs() -> List[WASpec]:
    out: List[WASpec] = []
    genders = ["M","F"] + (["X"] if WA_INCLUDE_MIXED else [])
    for gender in genders:
        for pool in ("LCM","SCM"):
            out.append(WASpec("WR", pool, gender))
            out.append(WASpec("WJ", pool, gender))
            out.append(WASpec("CR_AMERICAS", pool, gender))
        out.append(WASpec("OR","LCM", gender))
    return out

def wa_download_xlsx(page, url: str, out_dir: str) -> str:
    page.goto(url, wait_until="networkidle", timeout=120_000)
    try:
        page.get_by_role("button", name=re.compile(r"Accept Cookies|Accept all", re.I)).click(timeout=2500)
    except Exception:
        pass

    with page.expect_download(timeout=120_000) as dl_info:
        try:
            page.get_by_role("link", name=re.compile(r"\bXLSX\b", re.I)).click(timeout=10_000)
        except Exception:
            page.get_by_role("link", name=re.compile(r"Download", re.I)).click(timeout=10_000)

    download = dl_info.value
    filename = download.suggested_filename or f"wa_{uuid.uuid4().hex}.xlsx"
    path = os.path.join(out_dir, filename)
    download.save_as(path)
    return path

def wa_parse_xlsx(xlsx_path: str) -> List[Dict[str, Any]]:
    wb = load_workbook(xlsx_path, data_only=True)
    rows_out: List[Dict[str, Any]] = []

    def norm(v: Any) -> str:
        return _strip(v)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        values = list(ws.values)
        if not values:
            continue

        header_idx = None
        header_map: Dict[str, int] = {}

        for i, row in enumerate(values[:60]):
            row_norm = [norm(x).lower() for x in row]
            if any("event" in c for c in row_norm) and any("time" in c for c in row_norm):
                header_idx = i
                for j, c in enumerate(row_norm):
                    if "event" in c:
                        header_map["event"] = j
                    elif "time" in c:
                        header_map["time"] = j
                    elif "athlete" in c or "swimmer" in c or "record holder" in c:
                        header_map["athlete"] = j
                    elif "country" in c or "nation" in c:
                        header_map["country"] = j
                    elif "date" in c:
                        header_map["date"] = j
                    elif "place" in c or "location" in c or "venue" in c:
                        header_map["location"] = j
                    elif "competition" in c or "meet" in c:
                        header_map["competition"] = j
                break

        if header_idx is None or "event" not in header_map or "time" not in header_map:
            continue

        for row in values[header_idx+1:]:
            event = norm(row[header_map["event"]])
            t = norm(row[header_map["time"]])
            if not event or not t:
                continue

            athlete = norm(row[header_map.get("athlete", -1)]) if "athlete" in header_map else ""
            country = norm(row[header_map.get("country", -1)]) if "country" in header_map else ""
            date_raw = row[header_map.get("date", -1)] if "date" in header_map else ""
            date_str = parse_date(date_raw) or None

            # ✅ FIX: location SIEMPRE inicializada
            location = norm(row[header_map.get("location", -1)]) if "location" in header_map else ""
            competition = norm(row[header_map.get("competition", -1)]) if "competition" in header_map else ""

            rows_out.append({
                "event": event,
                "time": t,
                "athlete": athlete,
                "country": country,
                "date": date_str,
                "location": location,
                "competition": competition,
            })

    return rows_out

# -------------------------- Wikipedia parsers --------------------------

def http_get(url: str, timeout: int = 30) -> str:
    r = requests.get(url, timeout=timeout, headers={"User-Agent": HTTP_UA})
    r.raise_for_status()
    return r.text

def wiki_table_context(table) -> str:
    """
    Devuelve un contexto combinado para inferir género/piscina aunque la tabla tenga <caption>.
    Wikipedia a menudo pone el género en el heading (h2/h3/h4) y el caption es genérico.
    """
    parts = []

    cap = table.find("caption")
    if cap:
        parts.append(cap.get_text(" ", strip=True))

    # Captura hasta 2 headings previos (p.ej. "Piscina larga" + "Masculino")
    prev = table.find_previous(["h4", "h3", "h2"])
    if prev:
        parts.append(prev.get_text(" ", strip=True))
        prev2 = prev.find_previous(["h4", "h3", "h2"])
        if prev2:
            parts.append(prev2.get_text(" ", strip=True))

    return " | ".join([p for p in parts if p]).lower()


def wiki_guess_gender(ctx: str) -> Optional[str]:
    # Masculino
    if any(x in ctx for x in ["hombres", "hombre", "varones", "masculino", "men", "male", "boys"]):
        return "M"
    # Femenino
    if any(x in ctx for x in ["mujeres", "mujer", "damas", "femenino", "women", "female", "girls"]):
        return "F"
    # Mixto (si existiera)
    if any(x in ctx for x in ["mixto", "mixed"]):
        return "X"
    return None


def wiki_guess_pool(ctx: str) -> Optional[str]:
    ctx = (ctx or "").lower()
    if any(x in ctx for x in ["piscina corta","pileta corta","short course","scm","25"]):
        return "SCM"
    if any(x in ctx for x in ["piscina larga","pileta larga","long course","lcm","50"]):
        return "LCM"
    return None


def wiki_parse_records(url: str, default_pool: str = "LCM", default_gender: Optional[str] = None) -> List[Dict[str, Any]]:
    html = http_get(url, timeout=40)
    soup = BeautifulSoup(html, "html.parser")

    out: List[Dict[str, Any]] = []
    tables = soup.find_all("table", class_=re.compile("wikitable"))
    for t in tables:
        ctx = wiki_table_context(t)
        pool = wiki_guess_pool(ctx) or default_pool
        gender = wiki_guess_gender(ctx) or default_gender

        rows = t.find_all("tr")
        if not rows:
            continue
        head_cells = [c.get_text(" ", strip=True).lower() for c in rows[0].find_all(["th","td"])]

        def find_col(keys: Iterable[str]) -> Optional[int]:
            for i, c in enumerate(head_cells):
                for k in keys:
                    if k in c:
                        return i
            return None

        c_event = find_col(["event","prueba"])
        c_time  = find_col(["time","marca","tiempo"])
        c_swim  = find_col(["swimmer","record holder","nadador","athlete"])
        c_nat   = find_col(["nation","country","país","pais"])
        c_date  = find_col(["date","fecha"])
        c_meet  = find_col(["meet","competition","competición","competicion"])
        c_loc   = find_col(["location","place","lugar","venue"])

        if c_event is None or c_time is None:
            continue

        for r in rows[1:]:
            cells = [c.get_text(" ", strip=True) for c in r.find_all(["th","td"])]
            if len(cells) <= max(c_event, c_time):
                continue

            ev = cells[c_event]
            tm = cells[c_time]
            ms = parse_time_to_ms(tm)
            if ms is None:
                continue

            dist, stroke, type_probe = parse_event(ev)
            if dist is None or stroke is None:
                continue

            swimmer = cells[c_swim] if (c_swim is not None and c_swim < len(cells)) else ""
            nat = cells[c_nat] if (c_nat is not None and c_nat < len(cells)) else ""
            d_raw = cells[c_date] if (c_date is not None and c_date < len(cells)) else ""
            d_iso = parse_date(d_raw) or None
            meet = cells[c_meet] if (c_meet is not None and c_meet < len(cells)) else ""
            loc = cells[c_loc] if (c_loc is not None and c_loc < len(cells)) else ""

            out.append({
                "pool": pool,
                "gender": gender,
                "event": ev,
                "distance": dist,
                "stroke": stroke,
                "type_probe": type_probe,
                "time_ms": ms,
                "athlete": swimmer,
                "athlete_country": nat,              # atleta
                "record_date": d_iso,
                "competition": meet,
                "competition_location": loc,         # ✅ lugar del evento
                "source_url": url,
                "source_name": "Wikipedia",
                "source_note": "WIKI",
            })

    return out

# -------------------------- Payload builder --------------------------

def build_payload(
    record_scope: str,
    record_type: str,
    pool: str,
    gender: str,
    distance: int,
    stroke: str,
    time_ms: int,
    athlete_name: str,
    athlete_country: str,
    record_date: str,
    competition_name: str,
    competition_location: str,
    source_name: str,
    source_url: str,
    source_note: str,
    type_probe: str,
) -> Dict[str, Any]:
    t2 = format_ms_to_hms_2dp(int(time_ms))
    return {
        "gender": gender_label(gender),
        "category": "Open",
        "pool_length": pool_label(pool),
        "stroke": stroke,
        "distance": int(distance),
        "time_ms": int(time_ms),
        "time_clock_2dp": t2,
        "time_clock": t2,
        "record_scope": record_scope,
        "record_type": record_type,
        "competition_name": competition_name or "",
        "competition_location": competition_location or "",  # ✅ NUEVO
        "athlete_name": athlete_name or "",
        "country": athlete_country or "",  # atleta
        "city": "",                        # atleta (no suele venir)
        "record_date": parse_date(record_date) or None,
        "last_updated": RUN_DATE,
        "source_name": source_name or "",
        "source_url": source_url or "",
        "source_note": source_note or "",
        "type_probe": (type_probe or "individual"),
        "is_active": True,
    }

# -------------------------- Runners --------------------------

def run_wa(sb: SB) -> Dict[str, int]:
    stats = {"seen": 0, "updated": 0, "inserted": 0, "filled": 0, "unchanged": 0, "skipped": 0, "errors": 0}
    tmp_dir = f"/tmp/mdv_wa_{RUN_ID}"
    os.makedirs(tmp_dir, exist_ok=True)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()

        for spec in wa_specs():
            url = wa_url(spec)
            print(f"🔎 WA | {spec.code} | {spec.pool} | {spec.gender} | {url}")
            try:
                xlsx_path = wa_download_xlsx(page, url, tmp_dir)
                rows = wa_parse_xlsx(xlsx_path)

                record_scope, record_type = wa_scope_and_type(spec.code, spec.pool)
                seen_keys = set()

                for r in rows:
                    dist, stroke, type_probe = parse_event(r.get("event", ""))
                    ms = parse_time_to_ms(r.get("time", ""))
                    if dist is None or stroke is None or ms is None:
                        stats["skipped"] += 1
                        continue

                    # ✅ dedup por clave (evita dobles inserts y 23505 dentro de la corrida)
                    key = (spec.gender, pool_label(spec.pool), dist, stroke, record_type, record_scope)
                    if key in seen_keys:
                        continue
                    seen_keys.add(key)

                    payload = build_payload(
                        record_scope=record_scope,
                        record_type=record_type,
                        pool=spec.pool,
                        gender=spec.gender,
                        distance=dist,
                        stroke=stroke,
                        time_ms=ms,
                        athlete_name=r.get("athlete", ""),
                        athlete_country=r.get("country", ""),
                        record_date=r.get("date", ""),
                        competition_name=r.get("competition", ""),
                        competition_location=r.get("location", ""),  # ✅ lugar del evento
                        source_name="World Aquatics",
                        source_url=url,
                        source_note="XLSX",
                        type_probe=type_probe,
                    )

                    stats["seen"] += 1
                    try:
                        status = sb.upsert_record(payload)
                        if status == "inserted":
                            stats["inserted"] += 1
                        elif status == "updated":
                            stats["updated"] += 1
                        elif status == "filled":
                            stats["filled"] += 1
                        else:
                            stats["unchanged"] += 1
                    except Exception as e:
                        stats["errors"] += 1
                        print(f"❌ WA row error: {e}")
            except Exception as e:
                stats["errors"] += 1
                print(f"❌ WA {spec.code} {spec.pool} {spec.gender} error: {e}")
                continue

        browser.close()

    shutil.rmtree(tmp_dir, ignore_errors=True)
    return stats

def run_sudam(sb: SB) -> Dict[str, int]:
    stats = {"seen": 0, "updated": 0, "inserted": 0, "filled": 0, "unchanged": 0, "skipped": 0, "errors": 0}
    rows = wiki_parse_records(WIKI_SUDAM_URL, default_pool="LCM", default_gender=None)
    print(f"🌎 SUDAM source=WIKI filas={len(rows)}")

    for r in rows:
        g = r.get("gender")
        if g not in ("M","F"):
            stats["skipped"] += 1
            continue

        record_scope = "Sudamericano"
        record_type = "Récord Sudamericano" + (" SC" if pool_label(r["pool"]) == "SCM" else "")

        payload = build_payload(
            record_scope=record_scope,
            record_type=record_type,
            pool=r["pool"],
            gender=g,
            distance=r["distance"],
            stroke=r["stroke"],
            time_ms=r["time_ms"],
            athlete_name=r.get("athlete",""),
            athlete_country=r.get("athlete_country",""),
            record_date=r.get("record_date",""),
            competition_name=r.get("competition",""),
            competition_location=r.get("competition_location",""),
            source_name=r.get("source_name","Wikipedia"),
            source_url=r.get("source_url",WIKI_SUDAM_URL),
            source_note=r.get("source_note","WIKI"),
            type_probe=r.get("type_probe","individual"),
        )

        stats["seen"] += 1
        try:
            status = sb.upsert_record(payload)
            if status == "inserted":
                stats["inserted"] += 1
            elif status == "updated":
                stats["updated"] += 1
            elif status == "filled":
                stats["filled"] += 1
            else:
                stats["unchanged"] += 1
        except Exception as e:
            stats["errors"] += 1
            print(f"❌ SUDAM row error: {e}")

    return stats

def run_panam_games(sb: SB) -> Dict[str, int]:
    stats = {"seen": 0, "updated": 0, "inserted": 0, "filled": 0, "unchanged": 0, "skipped": 0, "errors": 0}
    rows = wiki_parse_records(WIKI_PANAM_GAMES_URL, default_pool="LCM", default_gender=None)
    print(f"🌎 PANAM_GAMES source=WIKI filas={len(rows)}")

    for r in rows:
        g = r.get("gender")
        if g not in ("M","F"):
            stats["skipped"] += 1
            continue

        record_scope = "Panamericano"
        record_type = "Récord Panamericano" + (" SC" if pool_label(r["pool"]) == "SCM" else "")

        payload = build_payload(
            record_scope=record_scope,
            record_type=record_type,
            pool=r["pool"],
            gender=g,
            distance=r["distance"],
            stroke=r["stroke"],
            time_ms=r["time_ms"],
            athlete_name=r.get("athlete",""),
            athlete_country=r.get("athlete_country",""),
            record_date=r.get("record_date",""),
            competition_name=r.get("competition",""),
            competition_location=r.get("competition_location",""),
            source_name=r.get("source_name","Wikipedia"),
            source_url=r.get("source_url",WIKI_PANAM_GAMES_URL),
            source_note=r.get("source_note","WIKI"),
            type_probe=r.get("type_probe","individual"),
        )

        stats["seen"] += 1
        try:
            status = sb.upsert_record(payload)
            if status == "inserted":
                stats["inserted"] += 1
            elif status == "updated":
                stats["updated"] += 1
            elif status == "filled":
                stats["filled"] += 1
            else:
                stats["unchanged"] += 1
        except Exception as e:
            stats["errors"] += 1
            print(f"❌ PANAM_GAMES row error: {e}")

    return stats

def run_arg_records(sb: SB) -> Dict[str, int]:
    """
    Récords Argentinos (absolutos) — módulo inicial.
    Fuente por defecto: Wikipedia (es) "Récords argentinos absolutos de natación".

    ✅ Pensado para ser reemplazable/extensible: si definís ARG_URLS (CSV) en env,
       se usarán esas URLs en lugar de la default.

    Nota: por ahora se carga como category="Open" (ver build_payload). Cuando incorporemos
    récords "Nacionales de Categoría" vamos a extender build_payload para aceptar category.
    """
    stats = {"seen": 0, "updated": 0, "inserted": 0, "filled": 0, "unchanged": 0, "skipped": 0, "errors": 0}

    env_urls = [u.strip() for u in os.getenv("ARG_URLS", "").split(",") if u.strip()]
    urls = env_urls or WIKI_ARG_URLS

    total_rows = 0
    for url in urls:
        try:
            rows = wiki_parse_records(url, default_pool="LCM", default_gender=None)
            if not rows:
                print(f"⚠️ ARG WIKI sin tablas/filas parseables | {url}")
                continue
            total_rows += len(rows)
            print(f"🇦🇷 ARG source=WIKI filas={len(rows)} | {url}")
        except Exception as e:
            print(f"❌ ARG ERROR parse wiki: {url} | {e}")
            stats["errors"] += 1
            continue

        for r in rows:
            g = r.get("gender")
            if g not in ("M", "F"):
                stats["skipped"] += 1
                continue

            # Etiquetas "human friendly" (alineado con SUDAM/PANAM/WA)
            record_scope = "Argentina"
            record_type = "Récord Argentino" + (" SC" if pool_label(r.get("pool","LCM")) == "SCM" else "")

            payload = build_payload(
                record_scope=record_scope,
                record_type=record_type,
                pool=r.get("pool","LCM"),
                gender=g,
                distance=int(r.get("distance", 0) or 0),
                stroke=r.get("stroke",""),
                time_ms=int(r.get("time_ms", 0) or 0),
                athlete_name=r.get("athlete",""),
                athlete_country=r.get("athlete_country",""),
                record_date=r.get("record_date",""),
                competition_name=r.get("competition",""),
                competition_location=r.get("competition_location",""),
                source_name=r.get("source_name","Wikipedia"),
                source_url=r.get("source_url", url),
                source_note=r.get("source_note","WIKI"),
                type_probe=r.get("type_probe","individual"),
            )

            # Validación mínima: distance y stroke tienen que existir
            if not payload.get("distance") or not payload.get("stroke") or not payload.get("time_ms"):
                stats["skipped"] += 1
                continue

            stats["seen"] += 1
            try:
                status = sb.upsert_record(payload)
                if status == "inserted":
                    stats["inserted"] += 1
                elif status == "updated":
                    stats["updated"] += 1
                elif status == "filled":
                    stats["filled"] += 1
                elif status == "unchanged":
                    stats["unchanged"] += 1
                else:
                    stats["skipped"] += 1
            except Exception as e:
                stats["errors"] += 1
                print(f"❌ ARG ERROR upsert {payload.get('stroke')} {payload.get('distance')} {g} {payload.get('pool_length')} | {e}")

    # Si parseamos tablas pero no insertamos nada, queda registrado en stats (seen=0 etc)
    return stats

def main() -> int:
    sb = SB(SUPABASE_URL, SUPABASE_KEY)

    print(f"MDV_UPDATER_VERSION={MDV_UPDATER_VERSION}")
    print(f"RUN_ID={RUN_ID}")
    print(f"Timestamp (UTC)={RUN_TS}")

    all_stats: Dict[str, Dict[str, int]] = {}
    all_stats["WA"] = run_wa(sb)
    all_stats["SUDAM"] = run_sudam(sb)
    all_stats["PANAM_GAMES"] = run_panam_games(sb)
    all_stats["ARG"] = run_arg_records(sb)

    print(f"Version: {MDV_UPDATER_VERSION}")
    print(f"Run ID: {RUN_ID}")
    print(f"Timestamp (UTC): {RUN_TS}")
    for k, st in all_stats.items():
        print(f"[{k}] seen={st['seen']} | inserted={st['inserted']} | updated={st['updated']} | filled={st['filled']} | unchanged={st['unchanged']} | skipped={st['skipped']} | errors={st['errors']}")

    # Política de fallo
    fatal_reasons: List[str] = []

    # Regla histórica (V15): si WA no procesa nada y encima hubo errores, el run es inválido.
    if all_stats["WA"]["seen"] == 0 and all_stats["WA"]["errors"] > 0:
        fatal_reasons.append(f"WA seen=0 errors={all_stats['WA']['errors']}")

    # Modo estricto (opcional): fail si cualquier fuente trae 0 filas o reporta errores.
    if MDV_STRICT:
        for k, st in all_stats.items():
            if st.get("seen", 0) == 0:
                fatal_reasons.append(f"{k} seen=0")
            if st.get("errors", 0) > 0:
                fatal_reasons.append(f"{k} errors={st['errors']}")

    if fatal_reasons:
        print("FATAL: " + " | ".join(fatal_reasons))
        return 1
    return 0

if __name__ == "__main__":
    sys.exit(main())
